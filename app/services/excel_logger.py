import os
from openpyxl import Workbook, load_workbook
from app.constants import EXCEL_FILE


def log_to_excel(post_id, platform, status, url):
    os.makedirs("output", exist_ok=True)

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Post ID", "Platform", "Status", "URL"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([post_id, platform, status, url])
    wb.save(EXCEL_FILE)